import requests
import xml.etree.ElementTree as ET
from datetime import datetime
import time
import os
import glob

FEED_URL_1 = "https://dsnkeyms.hook-dsn.pp.ua/feed/full-stock/catalog?token=c4b11a00d00657fffe4371ef43c3ea5f5d531cbd0903e8c1&offer_id=legacy"
FEED_URL_2 = "https://pkkopt.com.ua/content/export/e60e78a6b01d00d09fe25c1b666cc415.xml?1777544130"
OUTPUT_FILE = "prom_full_auto_feed.xml"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

def load_xml(url, name):
    print(f"--- DOWNLOAD {name} ---", flush=True)
    res = requests.get(url, headers=HEADERS, timeout=60)
    res.raise_for_status()
    return ET.fromstring(res.content)

# Розумний пошук Excel файлу та збір бази реальних ID з Прому
id_map = {}
try:
    import openpyxl
    excel_files = glob.glob("*.xlsx")
    if excel_files:
        excel_path = excel_files[0]
        print(f"--- READING EXCEL PARTNER KEY: {excel_path} ---", flush=True)
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet = wb.active
        
        # Шукаємо колонки з ID та Артикулом
        id_col, art_col = None, None
        for row in sheet.iter_rows(max_row=3, values_only=False):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip().lower()
                    if "ідентифікатор" in val or "id" in val:
                        id_col = cell.column
                    if "артикул" in val or "код_товару" in val or "код товару" in val:
                        art_col = cell.column
            if id_col and art_col:
                break
                
        if id_col and art_col:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= max(id_col, art_col):
                    p_id = row[id_col-1]
                    p_art = row[art_col-1]
                    if p_id and p_art:
                        id_map[str(p_art).strip()] = str(p_id).strip()
            print(f"--- SUCCESSFULLY MAPPED {len(id_map)} SEO PRODUCTS ---", flush=True)
        else:
            print("--- WARNING: COULD NOT FIND CORRECT COLUMNS IN EXCEL ---", flush=True)
    else:
        print("--- NO EXCEL FILE FOUND YET, USING DEFAULT ID RULES ---", flush=True)
except Exception as e:
    print(f"--- EXCEL READER NOTICE: {e} ---", flush=True)

try:
    root_out = ET.Element("yml_catalog", date=datetime.now().strftime("%Y-%m-%d %H:%M"))
    shop_out = ET.SubElement(root_out, "shop")
    categories_out = ET.SubElement(shop_out, "categories")
    offers_out = ET.SubElement(shop_out, "offers")

    # 1. ОБРОБКА DSN
    root1 = load_xml(FEED_URL_1, "DSN")
    for cat in root1.findall(".//category"):
        c_id = cat.get("id")
        p_id = cat.get("parentId")
        if c_id: cat.set("id", f"DSN_{c_id}")
        if p_id: cat.set("parentId", f"DSN_{p_id}")
        categories_out.append(cat)

    for offer in root1.findall(".//offer"):
        o_id = offer.get("id")
        if o_id: offer.set("id", f"DSN_{o_id}")
        c_id = offer.find("categoryId")
        if c_id is not None and c_id.text: c_id.text = f"DSN_{c_id.text}"
        v_code = offer.find("vendorCode")
        if v_code is not None and v_code.text: v_code.text = f"DSN_{v_code.text}"
        offers_out.append(offer)

    time.sleep(1)

    # 2. ОБРОБКА PKK
    root2 = load_xml(FEED_URL_2, "PKK")
    for cat in root2.findall(".//category"):
        c_id = cat.get("id")
        p_id = cat.get("parentId")
        if c_id and c_id.isdigit(): cat.set("id", str(int(c_id) + 900000000))
        if p_id and p_id.isdigit(): cat.set("parentId", str(int(p_id) + 900000000))
        categories_out.append(cat)

    group_data = {}
    all_pkk_offers = root2.findall(".//offer")

    for offer in all_pkk_offers:
        g_id = offer.get("group_id")
        if g_id:
            name = offer.find("name")
            description = offer.find("description")
            pictures = offer.findall("picture")
            if g_id not in group_data or (name is not None and description is not None):
                group_data[g_id] = {
                    "name": name.text if name is not None else None,
                    "description": description.text if description is not None else None,
                    "pictures": [p.text for p in pictures if p.text]
                }

    for offer in all_pkk_offers:
        o_id = offer.get("id")
        g_id = offer.get("group_id")
        
        # ПЕРЕВІРКА ЗА БАЗОЮ EXCEL ЕКСПОРТУ
        # Якщо цей товар вже є на Промі, підставляємо його точний рідний ID (24...)
        if o_id in id_map:
            offer.set("id", id_map[o_id])
        else:
            if o_id and o_id.isdigit():
                offer.set("id", o_id)

        if g_id and g_id.isdigit():
            numeric_group_id = int(g_id) + 900000000
            offer.set("group_id", str(numeric_group_id))
            
            if offer.find("name") is None or not offer.find("name").text:
                if group_data.get(g_id) and group_data[g_id]["name"]:
                    ET.SubElement(offer, "name").text = group_data[g_id]["name"]
            
            if offer.find("description") is None or not offer.find("description").text:
                if group_data.get(g_id) and group_data[g_id]["description"]:
                    ET.SubElement(offer, "description").text = group_data[g_id]["description"]
            
            if offer.find("picture") is None:
                if group_data.get(g_id) and group_data[g_id]["pictures"]:
                    for pic_url in group_data[g_id]["pictures"]:
                        ET.SubElement(offer, "picture").text = pic_url

            for p in offer.findall("param"):
                if p.get("name") in ["Цвет", "Колір"]:
                    offer.remove(p)
            ET.SubElement(offer, "param", name="Колір").text = f"№ {o_id}"

        c_id = offer.find("categoryId")
        if c_id is not None and c_id.text and c_id.text.isdigit():
            c_id.text = str(int(c_id.text) + 900000000)
        
        v_code = offer.find("vendorCode")
        if v_code is not None and v_code.text:
            v_code.text = v_code.text.strip()

        # Залишки
        quantity_tag = offer.find("quantity")
        is_available = offer.get("available")
        
        if is_available == "false" or (quantity_tag is not None and quantity_tag.text == "0"):
            offer.set("available", "false")
            if quantity_tag is not None: quantity_tag.text = "0"
            else: ET.SubElement(offer, "quantity").text = "0"
        elif quantity_tag is None or not quantity_tag.text:
            ET.SubElement(offer, "quantity").text = "5"

        price_tag = offer.find("price")
        oldprice_tag = offer.find("oldprice")
        if oldprice_tag is not None and oldprice_tag.text:
            if price_tag is not None: price_tag.text = oldprice_tag.text
            offer.remove(oldprice_tag)

        offers_out.append(offer)

    print("--- SAVING PERFECT CHRONO FEED ---", flush=True)
    tree = ET.ElementTree(root_out)
    tree.write(OUTPUT_FILE, encoding="utf-8", xml_declaration=True)
    print("--- SUCCESS ---", flush=True)

except Exception as e:
    print(f"ERROR: {e}", flush=True)
    exit(1)
